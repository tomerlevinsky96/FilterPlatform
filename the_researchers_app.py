import math
import os
from collections import defaultdict
from flask import Flask, render_template, request, redirect, url_for, flash, session
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask import Response
import pandas as pd
import io
from flask import Flask, render_template, request, jsonify, send_file
import psycopg2
from Configuration import DB_CONFIG,COLUMN_MAPPING
from psycopg2 import Error
import json
from openpyxl import Workbook
from io import BytesIO
import datetime
import re
from openpyxl.styles import Font, PatternFill
from search_values import search_values
from flask import request, abort
from werkzeug.serving import WSGIRequestHandler
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.serving import WSGIRequestHandler
from io import StringIO
import csv
from flask import Flask
from typing import List, Dict, Any
import sys
app = Flask(__name__, template_folder=os.path.abspath('Templates'))
app.secret_key = 'your_secret_key'  # Required for flash messaging

s1 = search_values()

# List of all available scan types
SCAN_TYPES = list(COLUMN_MAPPING.keys())

# Columns to exclude
EXCLUDE_COLUMNS = ['path', 'datetimescan']


def connect_to_db():
    try:
        connection = psycopg2.connect(**DB_CONFIG)
        connection.autocommit = True
        return connection
    except Error as e:
        print(f"Error while connecting to PostgreSQL: {e}")
        return None

###########################################filter2 functions
def flatten_values(value):
    if isinstance(value, tuple):
        return list(value)  # Convert tuple to list to extend it
    else:
        return [value]  # Return value as a list


def get_all_columns(connection, table_name):
    try:
        cursor = connection.cursor()
        query = f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'"
        cursor.execute(query)
        columns = [row[0] for row in cursor.fetchall() if row[0] not in EXCLUDE_COLUMNS]
        cursor.close()
        return columns
    except Error as e:
        print(f"Error fetching columns from {table_name}:", e)
        return []

def get_distinct_values(connection, app_column_name):
    try:
        cursor = connection.cursor()
        sql_column_name = COLUMN_MAPPING.get(app_column_name)
        if sql_column_name is None:
            print(f"No mapping found for column: {app_column_name}")
            return []

        query = f"SELECT DISTINCT {sql_column_name} FROM scans"
        cursor.execute(query)

        values = [row[0] for row in cursor.fetchall()]
        cursor.close()
        return values
    except Error as e:
        print(f"Error fetching distinct values for {app_column_name}:", e)
        return []

##############################filters1 functions
def clean_value(value):
    if value == 'NaT' or value == 'nan' or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, datetime.datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return value

def append_and_color_header(worksheet, headers, background_color):
    worksheet.append(headers)
    for cell in worksheet[worksheet.max_row]:
        cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
        cell.font = Font(color="000000")  # Black text
def process_flexible_data(headers, patient_code, data):
    # Create a defaultdict to store all values for each column
    data = sorted(data, key=lambda x: x[0])

    # Group the data by the first element (column index)
    grouped_data = defaultdict(list)
    for key, value in data:
        grouped_data[key].append(value)

    # Determine the maximum number of repetitions for any key
    max_repetitions = max(len(values) for values in grouped_data.values())

    # Create the output rows
    output_rows = [[patient_code] for _ in range(max_repetitions)]
    for key in sorted(grouped_data):
        values = grouped_data[key]
        for i in range(max_repetitions):
            if i < len(values):
                output_rows[i].append(values[i])
            else:
                output_rows[i].append('')

    return output_rows



############search_scans routes change to questionaire codes

@app.route('/get_filtered_patient_codes', methods=['POST'])
def get_filtered_patient_codes():
    print("Received request for filtered patient codes")

    # Extract data from form
    search_values.get_instance().set_start_date_of_scan(request.form.get('start_date_of_scan'))
    search_values.get_instance().set_end_date_of_scan(request.form.get('end_date_of_scan'))
    search_values.get_instance().set_start_hour_of_scan(request.form.get('start_hour_of_scan'))
    search_values.get_instance().set_end_hour_of_scan(request.form.get('end_hour_of_scan'))
    search_values.get_instance().set_study(request.form.get('study'))
    search_values.get_instance().set_group(request.form.get('Group'))
    search_values.get_instance().set_protocol(request.form.get('protocol'))
    search_values.get_instance().set_scan_number(request.form.get('scan_no'))
    search_values.get_instance().set_selected_genders(request.form.get('gender'))
    search_values.get_instance().set_age_from(request.form.get('age_from'))
    search_values.get_instance().set_age_to(request.form.get('age_to'))
    search_values.get_instance().set_height_from(request.form.get('height_from'))
    search_values.get_instance().set_height_to(request.form.get('height_to'))
    search_values.get_instance().set_weight_from(request.form.get('weight_from'))
    search_values.get_instance().set_weight_to(request.form.get('weight_to'))
    search_values.get_instance().set_dominant_hand(request.form.get('Dominant_hand'))
    search_values.get_instance().set_number_of_scans(request.form.get('number_of_scans'))

    # Extract protocol data
    scan_types = {scan_type: request.form.get(scan_type) for scan_type in SCAN_TYPES}
    search_values.get_instance().set_selected_types(scan_types)
    search_values.get_instance().set_connection(search_values.get_instance().connect_to_db())
    if search_values.get_instance().get_connection:
        search_values.get_instance().set_selected_patient_codes('')
        search_values.get_instance().set_dominant_hand_post('no')
        search_values.get_instance().set_update_subjects('yes')
        search_values.get_instance().append_to_data_output('NULL')
        patient_codes = search_values.get_instance().build_and_execute_query()
        patient_codes.sort()
        print("Returning patient codes:", patient_codes)
        return jsonify(patient_codes)
    return jsonify([])

@app.route('/get_groups_by_study', methods=['POST'])
def get_groups_by_study():
    study = request.form.get('study')
    if not study:
        return jsonify([])

    connection = connect_to_db()
    if connection:
        try:
            cursor = connection.cursor()
            # Query to get distinct groups for the selected study
            cursor.execute(
                "SELECT DISTINCT groupname FROM crf WHERE study = %s AND groupname IS NOT NULL AND groupname != 'nan' and groupname != '' and groupname != 'None'",
                (study,))
            groups = [row[0] for row in cursor.fetchall() if row[0] is not None]
            groups.sort()  # Sort alphabetically
            return jsonify(groups)
        except Error as e:
            print(f"Error fetching groups: {e}")
            return jsonify([])
        finally:
            connection.close()
    return jsonify([])

@app.route('/search_scans', methods=['GET', 'POST'])
def index():
    if session.get('authenticated'):
      connection = connect_to_db()
      if connection:
        cursor = connection.cursor()
        cursor.execute("SELECT distinct(questionairecode) FROM crf WHERE questionairecode IS NOT NULL AND questionairecode <> '' and questionairecode<>'nan'")
        patient_codes = [row[0] for row in cursor.fetchall() if row[0] is not None]
        patient_codes.sort()  # Sort patient codes alphabetically        # Sort patient codes alphabetically
        cursor.execute("SELECT DISTINCT Protocol FROM crf where Protocol <>'NULL' and Protocol<>'nan'")
        Protocols = [row[0] for row in cursor.fetchall() if row[0] is not None]
        Protocols.sort()  #
        cursor.execute("SELECT DISTINCT study FROM crf where study <>'NULL' and study<>'nan'")
        studies = [row[0] for row in cursor.fetchall() if row[0] is not None]
        studies.sort()
        cursor.execute("SELECT DISTINCT FLOOR(CAST(noscan AS NUMERIC))::INT AS noscan_int FROM crf WHERE noscan NOT IN ('NULL', 'nan');")
        scan_numbers = [row[0] for row in cursor.fetchall() if row[0] is not None]
        scan_numbers.sort()
        cursor.execute("select distinct(answer) from answers where questioneid='4' and answer <>'nan'")
        Dominant_hand = [row[0] for row in cursor.fetchall() if row[0] is not None]
        Dominant_hand.sort()
        cursor.execute("SELECT * FROM questiones WHERE questioneid >= 14 AND questioneid <= 15")
        custom_questions = [row[1] for row in cursor.fetchall() if row[0] is not None]
        cursor.execute("SELECT * FROM questiones WHERE questioneid >= 23 AND questioneid <= 28")
        education_work_questions = [row[1] for row in cursor.fetchall() if row[0] is not None]
        cursor.execute("SELECT * FROM questiones WHERE questioneid >= 313 AND questioneid <= 329")
        music_questions = [row[1] for row in cursor.fetchall() if row[0] is not None]
        all_questions =custom_questions + education_work_questions + music_questions
        connection.close()
      return render_template('search_scans.html', scan_types=SCAN_TYPES, patient_codes=patient_codes,studies=studies,protocols=Protocols,scan_numbers=scan_numbers,Dominant_hand=Dominant_hand)
    else:
        return render_template('loginPage.html')

def flatten_and_clean_values(result):
    """Flattens and cleans values in a single pass"""
    flattened = []
    for value in result:
        if isinstance(value, (list, tuple)):
            flattened.extend(clean_value(item) for item in value)
        else:
            flattened.append(clean_value(value))
    return flattened

@app.route('/export', methods=['POST'])
def export():
    while(1):
      try:
          search_values.get_instance().set_connection(search_values.get_instance().connect_to_db())
          cursor = search_values.get_instance().connection.cursor()
          file = request.files.get('file')
          selected_types = {}
          search_values.get_instance().Data_output = []
          # Populate selected_types dictionary
          for scan_type in SCAN_TYPES:
              values = request.form.getlist(scan_type)
              selected_types[scan_type] = values[0] if values else ''

          selected_genders = request.form.getlist('gender')[0] if request.form.getlist('gender') else ''

          # Set Singleton attributes

          search_values.get_instance().set_selected_types(selected_types)
          search_values.get_instance().set_dominant_hand_post(request.form.getlist('Dominant.hand'))
          if search_values.get_instance().get_dominant_hand_post():
              search_values.get_instance().append_to_data_output('answers.answer')
          search_values.get_instance().set_age_from(request.form.get('age_from'))
          search_values.get_instance().set_age_to(request.form.get('age_to'))
          search_values.get_instance().set_start_date_of_scan(request.form.get('start_date_of_scan'))
          search_values.get_instance().set_start_hour_of_scan(request.form.get('start_hour_of_scan'))
          search_values.get_instance().set_end_date_of_scan(request.form.get('end_date_of_scan'))
          search_values.get_instance().set_end_hour_of_scan(request.form.get('end_hour_of_scan'))
          search_values.get_instance().set_weight_from(request.form.get('weight_from'))
          search_values.get_instance().set_weight_to(request.form.get('weight_to'))
          search_values.get_instance().set_height_from(request.form.get('height_from'))
          search_values.get_instance().set_height_to(request.form.get('height_to'))
          search_values.get_instance().set_study(request.form.get('study'))
          search_values.get_instance().set_group(request.form.get('group'))
          search_values.get_instance().set_protocol(request.form.get('protocol'))
          search_values.get_instance().set_scan_number(request.form.get('scan_no'))
          search_values.get_instance().set_kepreppath(request.form.get('kepreppath'))
          search_values.get_instance().set_kepostpath(request.form.get('kepostpath'))
          search_values.get_instance().set_freesurferpath(request.form.get('freesurferpath'))
          #selected_patient_codes = request.form.getlist('selected_patient_codes')
          search_values.get_instance().set_dominant_hand(
              request.form.getlist('Dominant_hand')[0] if request.form.getlist('Dominant_hand') else ''
          )

          # Define the fields and process Data_output
          fields = ['Gender', 'crf.datetimescan', 'Ageofscan', 'weight', 'height', 'Study', 'Protocol', 'group',
                    'bidspath', 'resultspath', 'rawdatapath', 'Dominant.hand', 'kepreppath', 'kepostpath',
                    'freesurferpath']
          for field in fields:
              value = request.form.get(field)
              if value and value not in ['None']:
                  search_values.get_instance().append_to_data_output(value)

          # Database query and Excel export
          if search_values.get_instance().connection:
              wb = Workbook()
              ws = wb.active
              ws.title = "Subject details at the time of the scan"

              # Headers setup
              replacements = {
                  'answers.answer': 'dominant_hand',
                  'crf.scanid': 'scan_id',
                  'weight': 'weight(kg)',
                  'height': 'height(m)',
                  'preprocessedpath': 'keprep_path',
                  'postprocessedpath': 'kepost_path',
                  'freesurferpath': 'freesurfer_path',
                  'rawdatapath':'rawdata_path',
                  'bidspath':'bids_path',
                  'Ageofscan':'age_of_scan',
                  'Study':'study',
                  'Protocol':'protocol',
                  'Gender':'gender'
              }
              headers = ["scan_id","subject_code", "questionaire_code"] + [
                  "preprocessed path" if col == "kepreppath" else
                  "postprocessed path" if col == "kepostpath" else
                  "freesurfer path" if col == "freesurferpath" else

                  replacements.get(col, col)
                  for col in search_values.get_instance().get_data_output()
              ]
              append_and_color_header(ws, headers, "FFFFFF00")

              search_values.get_instance().set_update_subjects('no')
              search_values.get_instance().set_dominant_hand_post('yes')

              # Query and populate data rows
              patient_codes = json.loads(request.form.get('all_selected_patient_codes'))
              if len(patient_codes)==0 or len(search_values.get_instance().Data_output)==0:
                  break
              placeholders = ','.join(['%s'] * len(patient_codes))
              # Batch fetch all subject GUIDs
              query = f"SELECT scanid,guid,questionairecode FROM crf WHERE questionairecode IN ({placeholders})"
              cursor.execute(query, patient_codes)
              search_values.get_instance().set_selected_patient_codes(patient_codes)
              all_results = search_values.get_instance().build_and_execute_query()
              # Process results more efficiently
              rows_to_append = []
              for results in all_results:
                   rows_to_append.append(list(results))

              # Sort once at the end
              rows_to_append.sort(key=lambda x: x[1])
              size_in_bytes = sys.getsizeof(all_results)
              # Write rows to Excel using standard append method
              for row in rows_to_append:
                  ws.append(row)  # Using the standard append method instead of append_rows

              # Create and return Excel file
              excel_file = BytesIO()
              wb.save(excel_file)
              excel_file.seek(0)

              return send_file(
                  excel_file,
                  mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  as_attachment=True,
                  download_name='analysis_results.csv'
              )
              break
          return Response("No data found or an error occurred", status=400)

      except Exception as e:
        app.logger.error(f"Error processing export: {str(e)}")
        continue



########## filter1
# @app.route('/', methods=['GET', 'POST'])
# def filter_page():
#     # Example data for locations
#     locations = ["New York", "Los Angeles", "Chicago", "Miami"]
#
#     # Existing context data
#     return render_template('search_scans.html', all_questions=locations)

def are_codes_in_scanid_format(codes,scan_id_or_code):
    """
    Check if all codes in the array match the format yyyymmdd_hhmm.

    Args:
        codes (list): List of code strings to validate.

    Returns:
        bool: True if all codes match the format, False otherwise.
    """
    # Define the regex pattern for yyyymmdd_hhmm
    if scan_id_or_code=='scan_id':
       pattern = r'^\[?\d{8}_\d{4}\]?$'
    elif scan_id_or_code=='questionirecode':
       pattern = r'^\[?[^\s]*\]?$'
    for code in codes:
        if not re.match(pattern, code):
            return False
    return True




@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files and 'additionalDetails' not in request.form and 'selectedQuestions' not in request.form:
        return jsonify({'error': 'No file, data, or selected questions provided'})
    detail_type = request.form.get('detailType')
    search_values.get_instance().set_connection(search_values.get_instance().connect_to_db())
    file = request.files.get('file')
    data_text = request.form.get('additionalDetails')
    codes_or_scanids_array = []
    if file and file.filename:
        try:
            file_content = file.read()
            if file.filename.endswith('.csv'):
                df = pd.read_csv(io.StringIO(file_content.decode('utf-8')))
            elif file.filename.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(io.BytesIO(file_content))
            else:
                return jsonify({'error': 'Unsupported file format'})

            codes_or_scanids_array = df.values.flatten().tolist()
            # Regular expression pattern for matching paths or IDs
            pattern = re.compile(r'^([A-Za-z]:?(?:[\w-]+[/\\])*[\w-]+|[A-Za-z0-9_]+|[A-Za-z0-9]{6,})$')

            codes_or_scanids_array = [
                code for code in codes_or_scanids_array
                if (isinstance(code, str) and pattern.match(code)) or
                   (isinstance(code, (int, float)) and not pd.isna(code) and pattern.match(str(code)))
            ]
            if 'Uploaded' in codes_or_scanids_array:
                codes_or_scanids_array=codes_or_scanids_array[:-3]



        except Exception as e:
            return jsonify({'error': str(e)})

    elif data_text and len(codes_or_scanids_array)==0:
        codes_or_scanids_array.extend(data_text.split())

    elif not codes_or_scanids_array:
        return jsonify({'error': 'No valid patient codes provided'})

    conn = connect_to_db()
    if not conn:
        return jsonify({'error': 'Unable to connect to the database'})
    search_values.get_instance().Data_output = []
    search_values.get_instance().append_to_data_output('scans.bidspath')
    search_values.get_instance().append_to_data_output('scans.rawdatapath')
    search_values.get_instance().append_to_data_output('scans.kepreppath')
    search_values.get_instance().append_to_data_output('scans.kepostpath')
    search_values.get_instance().append_to_data_output('scans.freesurferpath')
    search_values.get_instance().set_update_subjects('no')
    search_values.get_instance().set_dominant_hand_post('yes')
    cursor = search_values.get_instance().connection.cursor()
    if detail_type == 'QuestionaireCode':
        search_values.get_instance().append_to_data_output('scanid')
        if are_codes_in_scanid_format(codes_or_scanids_array,'questionirecode')==False:
            return jsonify({'error': 'No valid questionaire codes provided'})
        codes_or_scanids_array = [code.strip('[]') for code in codes_or_scanids_array]

        if search_values.get_instance().connection:
            wb = Workbook()
            ws = wb.active
            ws.title = "Subject details at the time of the scan"

            # Headers setup
            replacements = {
                'answers.answer': 'Dominant hand',
                'crf.datetimescan': 'date time of scan',
                'weight': 'weight(kg)',
                'height': 'height(m)',
                'rawdatapath': 'rawdata_path',
                'kepreppath':'keprep_path',
                'kepostpath':'kepost_path',
                'bidspath':'bids_path',
                'freesurferpath': 'freesurfer_path'
            }
            headers = ["scan_ids","gui","questionairecode_code"] + [replacements.get(col, col) for col in
                                                               search_values.get_instance().get_data_output()]
            headers.pop()
            append_and_color_header(ws, headers, "FFFFFF00")
            search_values.get_instance().set_update_subjects('no')
            search_values.get_instance().set_dominant_hand_post('yes')
            # Query and populate data rows
            rows_to_append = []
            search_values.get_instance().set_additinal_information('yes')
            search_values.get_instance().set_selected_patient_codes(codes_or_scanids_array)
            results = search_values.get_instance().build_and_execute_query()
            if results:
               for result in results:
                   cleaned_result = [item for value in result for item in flatten_values(clean_value(value))]
                   rows_to_append.append(cleaned_result)
            else:
                return Response("No data found or an error occurred", status=400)

            rows_to_append.sort(key=lambda x: x[0])
            for row in rows_to_append:
                row.pop()
            for row in rows_to_append:
                ws.append(row)

            # Create an Excel file in memory
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            search_values.get_instance().set_additinal_information('no')
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='path_results.csv'
            )

        return Response("No data found or an error occurred", status=400)

    elif detail_type == 'scanid':
        if search_values.get_instance().connection:
            wb = Workbook()
            ws = wb.active
            ws.title = "Subject details at the time of the scan"
            # Headers setup
            replacements = {
                'answers.answer': 'Dominant hand',
                'crf.datetimescan': 'date time of scan',
                'weight': 'weight(kg)',
                'height': 'height(m)',
                'kepreppath': 'keprep path',
                'kepostpath': 'kepost path',
                'freesurferpath': 'freesurfer path'
            }
            headers = ["scan_ids"]+["guid"]+["questionaire_code"]+ [replacements.get(col, col) for col in
                                                               search_values.get_instance().get_data_output()]
            append_and_color_header(ws, headers, "FFFFFF00")
            search_values.get_instance().set_update_subjects('no')
            search_values.get_instance().set_dominant_hand_post('yes')
            # Query and populate data rows
            rows_to_append = []
            if are_codes_in_scanid_format(codes_or_scanids_array,'scan_id')==False:
               return jsonify({'error': 'No valid scan ids provided'})
            codes_or_scanids_array = [code.strip('[]') for code in codes_or_scanids_array]
            search_values.get_instance().set_selected_scanids(codes_or_scanids_array)

            results = search_values.get_instance().build_and_execute_query()
            if results:
               for result in results:
                cleaned_result = [item for value in result for item in flatten_values(clean_value(value))]
                rows_to_append.append(cleaned_result)
            else:
                return Response("No data found or an error occurred", status=400)
            rows_to_append.sort(key=lambda x: x[0])
            for row in rows_to_append:
                ws.append(row)

            # Create an Excel file in memory
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='results.xlsx'
            )

        return Response("No data found or an error occurred", status=400)
@app.route('/get_questions', methods=['GET'])
def get_questions():
    category = request.args.get('category')
    category_index_ranges = {
        'דמוגרפי כללי': [(3, 13)],
        'שפה ושיוך': [(14, 17)],
        'מצב משפחתי': [(17, 22)],
        'השכלה ומקצוע': [(23, 28)],
        'תחביבים והעדפות': [(29, 32), (464, 481), (489, 499)],
        'אורח חיים ועמדות': [(33, 47), (299, 312), (461, 464), (484, 484)],
        'שאלון שינה': [(48, 73)],
        'מצב בריאותי': [(74, 207), (290, 298)],
        'שאלון אישיות': [(208, 255)],
        'שאלון פסיכומטרי': [(256, 264)],
        'שאלון חרדה': [(265, 269)],
        'שאלון פוביות': [(270, 277)],
        'שאלון צאצאים שורדי שואה': [(278, 289)],
        'שאלון מוזיקה': [(313, 329)],
        'שאלון תכנות': [(330, 339)],
        'שאלון סמארטפון': [(340, 368)],
        'שאלון דיכאון וחרדה': [(369, 386)],
        'שאלון פוסט טראומה': [(387, 434)],
        'שאלון שבעה באוקטובר': [(435, 460)],
        'שאלות סיום': [(482, 483)],
        'All the questions': [(3, 501)],
    }

    conn = connect_to_db()
    if not conn:
        return jsonify({'error': 'Unable to connect to the database'})

    try:
        with conn.cursor() as cur:
            if category == 'Most_common_questions':
                common_questions = [
                    ('gender', 'Gender (at the time of the scan)'),
                    ('datetimescan', 'Date and time of scan'),
                    ('Ageofscan', 'Ageofscan (at the time of the scan)'),
                    ('weight', 'Weight (kg) (at the time of the scan)'),
                    ('height', 'Height (cm) (at the time of the scan)'),
                    ('study', 'study'),
                    ('Protocol', 'Protocol'),
                    ('groupname', 'Group'),
                    ('4', 'Dominant hand'),
                    #('Scan Details', 'Scan Details')
                ]
                cur.execute("SELECT * FROM questiones WHERE questioneid >= 14 AND questioneid <= 15")
                custom_questions = cur.fetchall()
                cur.execute("SELECT * FROM questiones WHERE questioneid >= 23 AND questioneid <= 28")
                education_work_questions = cur.fetchall()
                cur.execute("SELECT * FROM questiones WHERE questioneid >= 313 AND questioneid <= 329")
                music_questions = cur.fetchall()

                all_questions = common_questions + custom_questions + education_work_questions + music_questions
                return jsonify({'questions': all_questions})
            elif category == 'subject_details_at_the_time_of_scan':
                patient_details = [
                    ('gender', 'Gender (at the time of the scan) '),
                    ('datetimescan', 'Date and time of scan'),
                    ('Ageofscan', 'Ageofscan (at the time of the scan)'),
                    ('weight', 'Weight (kg) (at the time of the scan)'),
                    ('height', 'Height (cm) (at the time of the scan)'),
                    ('study', 'study'),
                    ('Protocol', 'Protocol'),
                    ('groupname', 'Group'),
                    # ('Scan Details', 'Scan Details')
                ]
                return jsonify({'questions': patient_details})
            elif category in category_index_ranges:
                questions = []
                for start_index, end_index in category_index_ranges[category]:
                    cur.execute("SELECT * FROM questiones WHERE questioneid >= %s AND questioneid <= %s",
                                (start_index, end_index))
                    questions.extend(cur.fetchall())
                return jsonify({'questions': questions})
            else:
                return jsonify({'error': 'Invalid category'})
    except psycopg2.Error as e:
        print(f"Database error: {e}")
        return jsonify({'error': 'Database error occurred'})
    finally:
        if conn:
            conn.close()


@app.route('/')
def index2():
    session['authenticated'] = False
    return render_template('loginPage.html')

@app.route('/login', methods=['POST'])
def login():

    password = request.form['password']
    if password == "Bvn1123@":  # Example check
        session['authenticated'] = True
        return redirect(url_for('OptionPage'))
    else:
        flash('Incorrect password. Please try again.', 'error')
        return redirect(url_for('index'))

@app.route('/Get_additinal_information')
def filters1():
    if session.get('authenticated'):
        return render_template('Get_additinal_information.html')
    else:
        return render_template('loginPage.html')


@app.route('/HomePage')
def OptionPage():
    if session.get('authenticated'):
        return render_template('HomePage.html')
    else:
        return render_template('loginPage.html')

@app.route('/logout')
def logout():
    session.pop('authenticated', None)
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(host='0.0.0.0')
